from netmiko import ConnLogOnly
from datetime import datetime
from getpass import getpass
from dotenv import load_dotenv
import yaml
import logging
import os
from openpyxl import Workbook
from openpyxl import load_workbook
import smtplib
from email.mime.text import MIMEText
from email.mime.multipart import MIMEMultipart
from email.mime.base import MIMEBase
from email import encoders


# Load yaml file with all network devices
def load_devices(device_file="devices.yml"):
    device_dict = {}
    with open(device_file) as f:
        device_dict = yaml.safe_load(f)
    return device_dict


# create file name for session logging with timestamp
def get_log_name(device_name):
    log_name = "logs/session/" + get_time() + "_" + device_name + ".log"
    return log_name


# Get the current local date/time and format the object to a string in a readable format
def get_time():
    time = datetime.now()
    time = time.strftime("%Y-%m-%d_%Hh%Mm")
    return time


# Creat an xlsx file for the report
def create_xlsx(filename):
    wb = Workbook()
    ws = wb.create_sheet(title="General")
    header = ["Device", "Software", "Model", "Uptime"]
    ws.append(header)
    ws.auto_filter.ref = "A1:D1"
    ws = wb.create_sheet(title="Serial and Mac")
    header = ["Device", "Mac", "SN"]
    ws.append(header)
    ws.auto_filter.ref = "A1:C1"
    ws = wb.create_sheet(title="License")
    header = [
        "Device",
        "Smart Licensing",
        "Transport",
        "Last ACK",
        "Trust code installed",
    ]
    ws.append(header)
    ws.auto_filter.ref = "A1:C1"

    if "Sheet" in wb.sheetnames:
        std = wb["Sheet"]
        wb.remove(std)
    wb.save(filename)
    logging.info("LCM report file created in xlsx file type.")


# Write output to report
def write_output_xlsx(filename, output_list, device_name):
    version = "unknown"
    uptime = "unknown"
    mac_address = "unknown"
    sn = "unknown"
    model = "unknown"
    lic_smart_status = "unknown"
    lic_transport_type = "unknown"
    lic_last_ack = "unknown"
    lic_trust_code = "unknown"
    wb = load_workbook(filename)
    custom_sheet = wb["Serial and Mac"]

    for output in output_list:
        if output:
            if "version" in output:
                switches = output["version"]["switch_num"]
                version = output["version"]["xe_version"]
                model = output["version"]["chassis"]
                uptime = output["version"]["uptime"]
                for device_id in switches:
                    device_info = switches[device_id]
                    mac_address = device_info["mac_address"]
                    sn = device_info["system_sn"]
                    row_sr_mac = [device_name, mac_address, sn]
                    custom_sheet.append(row_sr_mac)
            elif "smart_licensing_status" in output:
                lic_smart_status = output["smart_licensing_status"][
                    "smart_licensing_using_policy"
                ]["status"]
                lic_transport_type = output["smart_licensing_status"]["transport"][
                    "type"
                ]
                lic_last_ack = output["smart_licensing_status"]["usage_reporting"][
                    "last_ack_received"
                ]

    row_general = [device_name, version, model, uptime]
    custom_sheet = wb["General"]
    custom_sheet.append(row_general)
    row_license = [
        device_name,
        lic_smart_status,
        lic_transport_type,
        lic_last_ack,
        lic_trust_code,
    ]
    custom_sheet = wb["License"]
    custom_sheet.append(row_license)
    wb.save(filename)
    logging.info("Output is processed and written to xlsx file.")


def send_command(net_connect, command):
    output = net_connect.send_command(command, use_genie=True)
    if isinstance(output, str):
        logging.error(
            f"Failed to get output from the command: <{command}>. Possibly the command syntax is wrong or unknown in the Cisco Genie library.",
            exc_info=True,
        )
        output = False
    return output


def delete_empty_log_files(file_path):
    if os.path.exists(file_path) and os.stat(file_path).st_size == 0:
        os.remove(file_path)


# Handle connection to network device, return boolean True when connection was succesfull and false if otherwhise
def connect_to_device(device, username, password, device_name, filename):
    log_file_name = get_log_name(device_name)
    device["username"] = username
    device["password"] = password
    device["session_log"] = log_file_name
    ip = device["host"]
    logging.info("*" * 60)
    logging.info(f"Starting connection to switch: {device_name} ({ip})")
    print(f"\nStarting connection to switch: {device_name} ({ip})")
    logging.info("*" * 60)
    net_connect = ConnLogOnly(**device)
    output_list = []
    cmd_list = ["show version", "show license all"]
    cmd_ok = True
    if net_connect:
        for cmd in cmd_list:
            output = send_command(net_connect, cmd)
            output_list.append(output)
            if not output:
                cmd_ok = False
        write_output_xlsx(filename, output_list, device_name)
        net_connect.disconnect()
        logging.info(f"Successfully closed connection to switch: {device_name} ({ip})")
        print(f"Successfully closed connection to switch: {device_name} ({ip})")
        if not cmd_ok:
            return False
        return True
    write_output_xlsx(filename, output_list, device_name)
    delete_empty_log_files(log_file_name)
    print(f"Couldn't connect to switch: {device_name} ({ip})")
    return False


# Prompt for username
def get_username():
    print("Username: ", end="")
    username = input()
    return username


def yes_or_no(prompt):
    valid_responses = {"y": True, "n": False}
    while True:
        # Convert the user's input to lowercase to standardize the comparison
        user_input = input(prompt).strip().lower()
        if user_input in valid_responses:
            return valid_responses[user_input]
        else:
            print("Invalid input. Please enter 'y' or 'n'.")


# Create message with how the script performed
def create_summary_message(devices_no_connect):
    print("")
    if devices_no_connect:
        email_body = "Couldn't connect or retrieve desired information from the following switche(s): "
        for device in devices_no_connect:
            email_body += f"\n  *{device}"
        email_body += (
            "\n\nCheck logging files in the /logs/ directory for more details."
        )
        print(email_body)
        print("\nReport has been generated and stored in the /reports/ directory.")

    else:
        email_body = (
            "\nConnecting and retrieving information from all switches was successfull."
        )
        # Command line output with result of script
        print(email_body)
        print("\nReport has been generated and stored in the /reports/ directory.")

    return email_body


# Send email with report as attachment
def send_email(filename, devices_no_connect, filename_logs, email_password):
    sender_email = os.getenv("SENDER_EMAIL")
    password_email = email_password
    receiver_email = os.getenv("RECEIVER_EMAIL")
    smtp_server = os.getenv("SMTP_SERVER")
    smtp_port = os.getenv("SMTP_PORT")
    email_body = create_summary_message(devices_no_connect)
    logging.info(email_body)

    # Create email
    message = MIMEMultipart()
    message["From"] = sender_email
    message["To"] = receiver_email
    message["Subject"] = "LCM Report Cisco Catalyst"
    message.attach(MIMEText(email_body, "plain"))

    # Attach file
    filenames = [filename]
    if devices_no_connect:
        filenames.append(filename_logs)
    for file_to_attach in filenames:
        part = MIMEBase("application", "octet-stream")
        try:
            with open(file_to_attach, "rb") as file:
                part.set_payload(file.read())
                encoders.encode_base64(part)
                part.add_header(
                    "Content-Disposition",
                    f'attachment; filename="{file_to_attach.split("/")[-1]}"',
                )
            message.attach(part)
            logging.info(f"Successfully added attachment {file_to_attach} to email.")
        except Exception:
            logging.error(f"Failed to attach {file_to_attach} to email.", exc_info=True)

    # Send email
    try:
        with smtplib.SMTP(smtp_server, smtp_port) as server:
            server.starttls()  # Upgrade the connection to TLS
            server.login(sender_email, password_email)
            server.sendmail(sender_email, receiver_email, message.as_string())
            logging.info("Email sent successfully with attachment.")
            print("Email was sent successfully with attachment.")
    except Exception:
        logging.error("Failed to send email", exc_info=True)
        print(
            "Failed to send email, check logging in /logs/detailed/ directory for more details."
        )


if __name__ == "__main__":
    # Get time now to calculate total runtime script in the end
    start_time = datetime.now()

    # Check if directories for logs and xlsx reports exists, if not create them
    os.makedirs("logs/detailed", exist_ok=True)
    os.makedirs("logs/session", exist_ok=True)
    os.makedirs("reports/", exist_ok=True)

    # Load environment variables, store username and passwords if present, otherwise prompt for input
    load_dotenv()
    username = (
        os.getenv("USERNAME_SSH") if os.getenv("USERNAME_SSH") else get_username()
    )
    password = os.getenv("PASSWORD_SSH") if os.getenv("PASSWORD_SSH") else getpass()

    # Check if an email needs to be generated or not
    email = yes_or_no(
        "Do you want to send an email with the results? Please enter 'y' or 'n': "
    )
    email_password = ""

    if email:
        if os.getenv("PASSWORD_EMAIL"):
            email_password = os.getenv("PASSWORD_EMAIL")
        else:
            print("Fill in the email password:")
            getpass()

    # Enable logging, put logging level on DEBUG if you want more detail
    filename_logs = f"logs/detailed/{get_time()}.log"
    logging.basicConfig(
        filename=filename_logs,
        level=logging.INFO,
        format="%(asctime)s %(levelname)s %(name)s %(message)s",
    )
    logger = logging.getLogger(__name__)

    # Set filename for the report and create it
    filename = "reports/LCM_REPORT_CISCO_CAT9_" + get_time() + ".xlsx"
    report = create_xlsx(filename)

    # Load all devices in a dictionary variable
    devices_dict = load_devices()

    # Store all cisco device hostnames in a list variable
    device_list = devices_dict["cisco"]

    # Connect to every switch and store information in xlsx file, store device names of all connections that failed
    devices_no_connect = []
    for device_name in device_list:
        device = devices_dict[device_name]
        connect = connect_to_device(device, username, password, device_name, filename)
        if not connect:
            devices_no_connect.append(device_name)

    # Send email with LCM report as attachment or only prompt results to screen
    if email:
        send_email(filename, devices_no_connect, filename_logs, email_password)
    else:
        create_summary_message(devices_no_connect)

    # Add total runtime script to logging
    end_time = datetime.now()
    logging.info(f"\n\nScript execution time: {end_time - start_time}\n\n")
